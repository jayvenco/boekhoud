from fastapi import APIRouter, Request, Form, Depends, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, func
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import IntegrityError
from backend.models.database import get_db
from backend.models.models import Income, IncomeCategory, IncomeReceipt
from backend.routers.auth import require_auth
from backend.services.files import save_receipts, delete_file, move_tmp_to_category
from backend.services.invoice_numbering import get_numbering_settings, get_next_invoice_number
from backend.services.fiscal_year import is_year_locked, get_locked_years
from backend.services.pdf_export import generate_incomes_pdf
from backend.services.i18n import t
from datetime import date, datetime
from typing import Optional, List
import csv
import io

router = APIRouter(prefix="/inkomsten")
templates = Jinja2Templates(directory="backend/templates")
templates.env.globals["t"] = t

# Standaardopties voor "Ontvangen op". Nieuwe optie toevoegen = één regel hier
# + één <option> in het formulier.
RECEIVED_VIA_OPTIONS = {
    "zakelijke_rekening": "Zakelijke rekening",
    "priverekening": "Privérekening",
    "creditcard": "Creditcard",
    "rekening_partner": "Rekening partner",
    "externe_praktijk": "Externe praktijk",
    "sumup": "SumUp",
    "payt": "Payt debiteurenbeheer",
    "infomedics": "Infomedics",
    "overig": "Overig",
}


def parse_date(d: str) -> date:
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(d.strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Ongeldige datum: {d}")


def validate_received_via(received_via: str, received_via_other: str):
    """Retourneert (received_via, received_via_other, foutmelding_of_None)."""
    if received_via not in RECEIVED_VIA_OPTIONS:
        return received_via, None, f"Onbekende optie voor 'Ontvangen op': {received_via}"
    other = received_via_other.strip() if received_via == "overig" else None
    if received_via == "overig" and not other:
        return received_via, other, "Specificeer op welke rekening of via welk betaalmiddel de inkomst is ontvangen."
    return received_via, other, None


def _apply_income_filters(query, q, from_date, to_date, category_id, year,
                          min_amount, max_amount, contact):
    if q:
        query = query.where(or_(
            Income.invoice_number.ilike(f"%{q}%"),
            Income.description.ilike(f"%{q}%"),
        ))
    if from_date:
        try:
            query = query.where(Income.date >= parse_date(from_date))
        except ValueError:
            pass
    if to_date:
        try:
            query = query.where(Income.date <= parse_date(to_date))
        except ValueError:
            pass
    if category_id:
        try:
            query = query.where(Income.category_id == int(category_id))
        except ValueError:
            pass
    if year:
        query = query.where(func.strftime("%Y", Income.date) == str(year))
    if min_amount:
        try:
            query = query.where(Income.amount >= float(min_amount))
        except ValueError:
            pass
    if max_amount:
        try:
            query = query.where(Income.amount <= float(max_amount))
        except ValueError:
            pass
    if contact:
        query = query.join(Income.category).where(or_(
            IncomeCategory.contact_firstname.ilike(f"%{contact}%"),
            IncomeCategory.contact_lastname.ilike(f"%{contact}%"),
            IncomeCategory.email.ilike(f"%{contact}%"),
            IncomeCategory.phone.ilike(f"%{contact}%"),
        ))
    return query


@router.get("", response_class=HTMLResponse)
async def list_incomes(
    request: Request, db: AsyncSession = Depends(get_db),
    q: str = "", from_date: str = "", to_date: str = "",
    category_id: str = "", year: str = "",
    min_amount: str = "", max_amount: str = "", contact: str = "",
):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    query = select(Income).options(
        selectinload(Income.category),
        selectinload(Income.receipts)
    ).order_by(Income.date.desc())
    query = _apply_income_filters(query, q, from_date, to_date, category_id,
                                  year, min_amount, max_amount, contact)
    result = await db.execute(query)
    incomes = result.scalars().all()
    cats = await db.execute(select(IncomeCategory))
    locked_years = await get_locked_years(db)
    filters = {"q": q, "from_date": from_date, "to_date": to_date,
               "category_id": category_id, "year": year,
               "min_amount": min_amount, "max_amount": max_amount, "contact": contact}
    return templates.TemplateResponse(request, "incomes/list.html", {
        "incomes": incomes, "categories": cats.scalars().all(),
        "received_via_labels": RECEIVED_VIA_OPTIONS,
        "locked_years": locked_years,
        "filters": filters,
        "active_filters": sum(1 for v in filters.values() if v),
    })


@router.get("/export/pdf")
async def export_incomes_pdf(
    request: Request, db: AsyncSession = Depends(get_db),
    q: str = "", from_date: str = "", to_date: str = "",
    category_id: str = "", year: str = "",
    min_amount: str = "", max_amount: str = "", contact: str = "",
):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    query = select(Income).options(
        selectinload(Income.category),
        selectinload(Income.receipts)
    ).order_by(Income.date.desc())
    query = _apply_income_filters(query, q, from_date, to_date, category_id,
                                  year, min_amount, max_amount, contact)
    result = await db.execute(query)
    incomes = result.scalars().all()

    parts = []
    if q:           parts.append(f"Zoeken: {q}")
    if from_date:   parts.append(f"Van: {from_date}")
    if to_date:     parts.append(f"Tot: {to_date}")
    if year:        parts.append(f"Boekjaar: {year}")
    if min_amount:  parts.append(f"Min: €{min_amount}")
    if max_amount:  parts.append(f"Max: €{max_amount}")
    if contact:     parts.append(f"Contact: {contact}")
    if category_id:
        cat_r = await db.get(IncomeCategory, int(category_id))
        if cat_r:   parts.append(f"Categorie: {cat_r.name}")
    filters_desc = " | ".join(parts)

    settings = request.state.settings
    company = settings.company_name if settings else ""
    buf = generate_incomes_pdf(incomes, company_name=company,
                               filters_desc=filters_desc,
                               received_via_labels=RECEIVED_VIA_OPTIONS)
    filename = f"inkomsten_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/nieuw", response_class=HTMLResponse)
async def new_income_form(request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    cats = await db.execute(select(IncomeCategory))
    numbering_settings = await get_numbering_settings(db)
    return templates.TemplateResponse(request, "incomes/form.html", {
        "categories": cats.scalars().all(), "income": None,
        "auto_numbering_enabled": numbering_settings.auto_enabled,
        "received_via_options": RECEIVED_VIA_OPTIONS
    })


@router.post("/nieuw")
async def create_income(
    request: Request,
    invoice_number: str = Form(...),
    category_id: int = Form(...),
    date_str: str = Form(..., alias="date"),
    amount: float = Form(...),
    description: str = Form(""),
    status: str = Form("niet_betaald"),
    received_via: str = Form(...),
    received_via_other: str = Form(""),
    receipts: List[UploadFile] = File(default=[]),
    ocr_tmp_file: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db)
):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user

    # Controleer of het boekjaar vergrendeld is
    income_year = parse_date(date_str).year
    if await is_year_locked(db, income_year):
        cats = await db.execute(select(IncomeCategory))
        return templates.TemplateResponse(request, "incomes/form.html", {
            "categories": cats.scalars().all(), "income": None,
            "received_via_options": RECEIVED_VIA_OPTIONS,
            "error": f"Boekjaar {income_year} is afgesloten. Inkomsten kunnen niet worden toegevoegd aan een afgesloten boekjaar."
        })

    received_via, received_via_other, via_error = validate_received_via(received_via, received_via_other)
    if via_error:
        cats = await db.execute(select(IncomeCategory))
        return templates.TemplateResponse(request, "incomes/form.html", {
            "categories": cats.scalars().all(), "income": None,
            "received_via_options": RECEIVED_VIA_OPTIONS,
            "error": via_error
        })

    # Check duplicate
    dup = await db.execute(select(Income).where(Income.invoice_number == invoice_number))
    if dup.scalar_one_or_none():
        cats = await db.execute(select(IncomeCategory))
        return templates.TemplateResponse(request, "incomes/form.html", {
            "categories": cats.scalars().all(), "income": None,
            "received_via_options": RECEIVED_VIA_OPTIONS,
            "error": f"Factuurnummer '{invoice_number}' bestaat al."
        })

    cat = await db.get(IncomeCategory, category_id)
    valid_files = [r for r in receipts if r and r.filename]
    saved = []
    if valid_files:
        saved = await save_receipts(valid_files, "inkomsten", cat.slug, invoice_number)
    elif ocr_tmp_file:
        p = await move_tmp_to_category(ocr_tmp_file, "inkomsten", cat.slug, invoice_number)
        if p:
            saved = [{"path": p, "suffix": None}]

    receipt_path = saved[0]["path"] if saved else None
    income = Income(
        invoice_number=invoice_number, category_id=category_id,
        date=parse_date(date_str), amount=amount, description=description,
        status=status, received_via=received_via, received_via_other=received_via_other,
        receipt_path=receipt_path
    )
    db.add(income)
    try:
        await db.flush()
    except IntegrityError:
        await db.rollback()
        cats = await db.execute(select(IncomeCategory))
        return templates.TemplateResponse(request, "incomes/form.html", {
            "categories": cats.scalars().all(), "income": None,
            "received_via_options": RECEIVED_VIA_OPTIONS,
            "error": f"Factuurnummer '{invoice_number}' is zojuist al gebruikt door een andere registratie. Vernieuw de pagina en probeer opnieuw."
        })
    for r in saved:
        db.add(IncomeReceipt(income_id=income.id, file_path=r["path"], suffix=r["suffix"]))
    await db.commit()
    return RedirectResponse("/inkomsten", status_code=302)


@router.get("/{id}/bewerken", response_class=HTMLResponse)
async def edit_income_form(id: int, request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    result = await db.execute(
        select(Income).options(selectinload(Income.category), selectinload(Income.receipts))
        .where(Income.id == id)
    )
    income = result.scalar_one_or_none()
    if not income:
        raise HTTPException(404)
    cats = await db.execute(select(IncomeCategory))
    year_locked = await is_year_locked(db, income.date.year)
    return templates.TemplateResponse(request, "incomes/form.html", {
        "categories": cats.scalars().all(), "income": income,
        "received_via_options": RECEIVED_VIA_OPTIONS,
        "year_locked": year_locked,
    })


@router.post("/{id}/bewerken")
async def update_income(
    id: int, request: Request,
    invoice_number: str = Form(...),
    category_id: int = Form(...),
    date_str: str = Form(..., alias="date"),
    amount: float = Form(...),
    description: str = Form(""),
    status: str = Form("niet_betaald"),
    received_via: str = Form(...),
    received_via_other: str = Form(""),
    receipts: List[UploadFile] = File(default=[]),
    ocr_tmp_file: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db)
):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    result = await db.execute(
        select(Income).options(selectinload(Income.receipts)).where(Income.id == id)
    )
    income = result.scalar_one_or_none()
    if not income:
        raise HTTPException(404)

    if await is_year_locked(db, income.date.year):
        cats = await db.execute(select(IncomeCategory))
        return templates.TemplateResponse(request, "incomes/form.html", {
            "categories": cats.scalars().all(), "income": income,
            "received_via_options": RECEIVED_VIA_OPTIONS,
            "year_locked": True,
            "error": f"Boekjaar {income.date.year} is afgesloten. Wijzigingen zijn niet toegestaan."
        })

    received_via, received_via_other, via_error = validate_received_via(received_via, received_via_other)
    if via_error:
        cats = await db.execute(select(IncomeCategory))
        result2 = await db.execute(
            select(Income).options(selectinload(Income.category), selectinload(Income.receipts)).where(Income.id == id)
        )
        return templates.TemplateResponse(request, "incomes/form.html", {
            "categories": cats.scalars().all(), "income": result2.scalar_one_or_none(),
            "received_via_options": RECEIVED_VIA_OPTIONS,
            "error": via_error
        })

    # Check duplicate (exclude self)
    dup = await db.execute(select(Income).where(Income.invoice_number == invoice_number, Income.id != id))
    if dup.scalar_one_or_none():
        cats = await db.execute(select(IncomeCategory))
        result2 = await db.execute(
            select(Income).options(selectinload(Income.category), selectinload(Income.receipts)).where(Income.id == id)
        )
        return templates.TemplateResponse(request, "incomes/form.html", {
            "categories": cats.scalars().all(), "income": result2.scalar_one_or_none(),
            "received_via_options": RECEIVED_VIA_OPTIONS,
            "error": f"Factuurnummer '{invoice_number}' bestaat al."
        })

    cat = await db.get(IncomeCategory, category_id)
    valid_files = [r for r in receipts if r and r.filename]
    if valid_files:
        new_saved = await save_receipts(valid_files, "inkomsten", cat.slug, invoice_number)
        for r in new_saved:
            db.add(IncomeReceipt(income_id=income.id, file_path=r["path"], suffix=r["suffix"]))
        if not income.receipt_path and new_saved:
            income.receipt_path = new_saved[0]["path"]

    income.invoice_number = invoice_number
    income.category_id = category_id
    income.date = parse_date(date_str)
    income.amount = amount
    income.description = description
    income.status = status
    income.received_via = received_via
    income.received_via_other = received_via_other
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        cats = await db.execute(select(IncomeCategory))
        result2 = await db.execute(
            select(Income).options(selectinload(Income.category), selectinload(Income.receipts)).where(Income.id == id)
        )
        return templates.TemplateResponse(request, "incomes/form.html", {
            "categories": cats.scalars().all(), "income": result2.scalar_one_or_none(),
            "received_via_options": RECEIVED_VIA_OPTIONS,
            "error": f"Factuurnummer '{invoice_number}' is zojuist al gebruikt door een andere registratie. Vernieuw de pagina en probeer opnieuw."
        })
    return RedirectResponse("/inkomsten", status_code=302)


@router.post("/{id}/bon-verwijderen/{receipt_id}")
async def delete_receipt(id: int, receipt_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    result = await db.execute(select(IncomeReceipt).where(IncomeReceipt.id == receipt_id, IncomeReceipt.income_id == id))
    receipt = result.scalar_one_or_none()
    if receipt:
        delete_file(receipt.file_path)
        await db.delete(receipt)
        await db.commit()
    return RedirectResponse(f"/inkomsten/{id}/bewerken", status_code=302)


@router.get("/import", response_class=HTMLResponse)
async def import_form(request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    return templates.TemplateResponse(request, "incomes/import.html", {"user": user})


@router.post("/import", response_class=HTMLResponse)
async def import_incomes(
    request: Request,
    bestand: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user

    filename = (bestand.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        return templates.TemplateResponse(request, "incomes/import.html", {
            "user": user,
            "fout": "Alleen CSV (.csv) en Excel (.xlsx) bestanden zijn toegestaan.",
        })

    content = await bestand.read()

    # ── Parse rows ────────────────────────────────────────────
    raw_rows: list[dict] = []
    parse_error: str | None = None

    if filename.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")  # strip BOM if present
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                raw_rows.append({k.strip().lower(): v.strip() for k, v in row.items()})
        except Exception as e:
            parse_error = f"CSV-leessfout: {e}"
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                raw_rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        except Exception as e:
            parse_error = f"Excel-leessfout: {e}"

    if parse_error:
        return templates.TemplateResponse(request, "incomes/import.html", {
            "user": user, "fout": parse_error,
        })

    # ── Load categories (name → object, case-insensitive) ─────
    cat_result = await db.execute(select(IncomeCategory))
    cat_map = {c.name.lower(): c for c in cat_result.scalars().all()}

    ns = await get_numbering_settings(db)
    locked_years: set[int] = await get_locked_years(db)

    imported, skipped = 0, []

    for row_num, row in enumerate(raw_rows, start=2):
        errors = []

        # datum
        raw_date = row.get("datum", "")
        try:
            record_date = parse_date(raw_date)
        except ValueError:
            errors.append(f"Ongeldige datum '{raw_date}'")
            record_date = None

        # bedrag
        raw_amount = row.get("bedrag", "").replace(",", ".")
        try:
            amount = float(raw_amount)
            if amount <= 0:
                raise ValueError
        except ValueError:
            errors.append(f"Ongeldig bedrag '{row.get('bedrag', '')}'")
            amount = None

        # categorie
        cat_name = row.get("categorie", "").lower()
        cat = cat_map.get(cat_name)
        if not cat:
            errors.append(f"Categorie '{row.get('categorie', '')}' bestaat niet")

        # boekjaar vergrendeld?
        if record_date and record_date.year in locked_years:
            errors.append(f"Boekjaar {record_date.year} is afgesloten")

        if errors:
            skipped.append({"rij": row_num, "reden": "; ".join(errors), "data": dict(row)})
            continue

        # factuurnummer
        invoice_number = row.get("factuurnummer", "").strip()
        if not invoice_number:
            if ns.auto_enabled:
                invoice_number = await get_next_invoice_number(db, record_date.year, "inkomsten", ns)
            else:
                skipped.append({"rij": row_num, "reden": "Geen factuurnummer opgegeven en auto-nummering is uitgeschakeld", "data": dict(row)})
                continue

        # duplicate check
        dup = await db.execute(select(Income).where(Income.invoice_number == invoice_number))
        if dup.scalar_one_or_none():
            skipped.append({"rij": row_num, "reden": f"Factuurnummer '{invoice_number}' bestaat al", "data": dict(row)})
            continue

        # status + ontvangen_op
        status_val = row.get("status", "niet_betaald").strip() or "niet_betaald"
        if status_val not in ("betaald", "niet_betaald"):
            status_val = "niet_betaald"

        received_via = row.get("ontvangen_op", "zakelijke_rekening").strip() or "zakelijke_rekening"
        if received_via not in RECEIVED_VIA_OPTIONS:
            received_via = "zakelijke_rekening"

        db.add(Income(
            invoice_number=invoice_number,
            category_id=cat.id,
            date=record_date,
            amount=amount,
            description=row.get("omschrijving", "").strip() or None,
            status=status_val,
            received_via=received_via,
        ))
        imported += 1

    await db.commit()

    return templates.TemplateResponse(request, "incomes/import.html", {
        "user": user,
        "imported": imported,
        "skipped": skipped,
        "total": len(raw_rows),
    })


@router.get("/import/voorbeeld-csv")
async def download_csv_example():
    return FileResponse(
        "backend/static/import_inkomsten_voorbeeld.csv",
        media_type="text/csv",
        filename="import_inkomsten_voorbeeld.csv",
    )


@router.get("/import/voorbeeld-xlsx")
async def download_xlsx_example():
    return FileResponse(
        "backend/static/import_inkomsten_voorbeeld.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="import_inkomsten_voorbeeld.xlsx",
    )


@router.post("/{id}/verwijderen")
async def delete_income(id: int, request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    result = await db.execute(
        select(Income).options(selectinload(Income.receipts)).where(Income.id == id)
    )
    income = result.scalar_one_or_none()
    if income and await is_year_locked(db, income.date.year):
        return RedirectResponse("/inkomsten?error=vergrendeld", status_code=302)
    if income:
        for r in income.receipts:
            delete_file(r.file_path)
        if income.receipt_path:
            delete_file(income.receipt_path)
        await db.delete(income)
        await db.commit()
    return RedirectResponse("/inkomsten", status_code=302)
