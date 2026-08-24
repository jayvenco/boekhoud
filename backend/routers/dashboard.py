from fastapi import APIRouter, Request, Depends, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from backend.models.database import get_db
from backend.models.models import Income, Expense, ExpenseCategory, IncomeCategory, Depreciation, ChecklistItem, YearClosure, FiscalYear, TimeEntry, HourCategory
from backend.routers.auth import require_auth
from backend.routers.checklist import get_checklist_summary
from backend.routers.hours import URENCRITERIUM_UREN
from backend.services.i18n import t
from backend.services.pdf_export import generate_yearly_pdf
from datetime import date, datetime
from typing import Optional
import csv
import io

router = APIRouter()
templates = Jinja2Templates(directory="backend/templates")
templates.env.globals["t"] = t


def year_filter(model, year: int):
    """SQLite-compatible year filter using strftime."""
    from sqlalchemy import func
    return func.strftime("%Y", model.date) == str(year)



def calc_depreciation_for_year(dep, year: int) -> float:
    start_year = dep.start_date.year
    end_year = start_year + dep.duration_years
    if year < start_year or year >= end_year:
        return 0.0
    residual = dep.residual_value or 0.0
    depreciable_base = dep.purchase_amount - residual
    return depreciable_base / dep.duration_years if dep.duration_years else 0.0


def calc_dep_overview(dep):
    residual = dep.residual_value or 0.0
    depreciable_base = dep.purchase_amount - residual
    annual = depreciable_base / dep.duration_years if dep.duration_years else 0.0
    today = date.today()
    years_elapsed = min(max(0, today.year - dep.start_date.year), dep.duration_years)
    depreciated = annual * years_elapsed
    book_value = max(dep.purchase_amount - depreciated, residual)
    end_year = dep.start_date.year + dep.duration_years

    schedule = []
    cumulative = 0
    for y in range(dep.start_date.year, end_year):
        cumulative += annual
        cumulative_capped = min(cumulative, depreciable_base)
        schedule.append({
            "year": y,
            "amount": annual,
            "cumulative": cumulative_capped,
            "book_value": max(dep.purchase_amount - cumulative_capped, residual),
            "is_current": y == today.year,
            "is_past": y < today.year,
        })

    return {
        "description": dep.expense.description or dep.expense.invoice_number,
        "invoice": dep.expense.invoice_number,
        "purchase": dep.purchase_amount,
        "residual": residual,
        "annual": annual,
        "depreciated": depreciated,
        "book_value": book_value,
        "start_year": dep.start_date.year,
        "end_year": end_year,
        "duration": dep.duration_years,
        "schedule": schedule,
    }


@router.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user

    year = datetime.now().year

    inc_result = await db.execute(
        select(func.sum(Income.amount)).where(year_filter(Income, year))
    )
    total_income = inc_result.scalar() or 0

    exp_result = await db.execute(
        select(func.sum(Expense.amount)).where(year_filter(Expense, year))
    )
    total_expenses = exp_result.scalar() or 0

    unpaid_result = await db.execute(
        select(func.sum(Income.amount)).where(
            Income.status == "niet_betaald",
            year_filter(Income, year)
        )
    )
    unpaid = unpaid_result.scalar() or 0

    recent_inc = await db.execute(
        select(Income).options(selectinload(Income.category))
        .order_by(Income.date.desc()).limit(5)
    )
    recent_exp = await db.execute(
        select(Expense).options(selectinload(Expense.category))
        .order_by(Expense.date.desc()).limit(5)
    )

    monthly_inc = await db.execute(
        select(
            func.strftime("%m", Income.date).label("month"),
            func.sum(Income.amount).label("total")
        ).where(year_filter(Income, year))
        .group_by(func.strftime("%m", Income.date))
        .order_by(func.strftime("%m", Income.date))
    )
    monthly_exp = await db.execute(
        select(
            func.strftime("%m", Expense.date).label("month"),
            func.sum(Expense.amount).label("total")
        ).where(year_filter(Expense, year))
        .group_by(func.strftime("%m", Expense.date))
        .order_by(func.strftime("%m", Expense.date))
    )

    inc_by_month = {int(r.month): float(r.total) for r in monthly_inc}
    exp_by_month = {int(r.month): float(r.total) for r in monthly_exp}
    months = ["Jan", "Feb", "Mrt", "Apr", "Mei", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Dec"]

    today = date.today()
    missing_receipt_result = await db.execute(
        select(Expense).options(selectinload(Expense.receipts)).where(
            Expense.is_recurring == True,
            Expense.date <= today,
            Expense.receipt_path.is_(None),
        )
    )
    missing_receipts = [e for e in missing_receipt_result.scalars().all() if not e.receipts]

    checklist_summary = await get_checklist_summary(db)

    return templates.TemplateResponse(request, "dashboard.html", {
        "total_income": total_income,
        "total_expenses": total_expenses,
        "profit": total_income - total_expenses,
        "unpaid": unpaid,
        "recent_incomes": recent_inc.scalars().all(),
        "recent_expenses": recent_exp.scalars().all(),
        "checklist_summary": checklist_summary,
        "missing_receipts_count": len(missing_receipts),
        "chart_months": months,
        "chart_income": [inc_by_month.get(i, 0) for i in range(1, 13)],
        "chart_expenses": [exp_by_month.get(i, 0) for i in range(1, 13)],
        "year": year,
    })


@router.get("/rapportages", response_class=HTMLResponse)
async def reports(request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user

    year = int(request.query_params.get("jaar", datetime.now().year))

    inc_by_cat = await db.execute(
        select(IncomeCategory.name, func.sum(Income.amount).label("total"))
        .join(Income, Income.category_id == IncomeCategory.id)
        .where(year_filter(Income, year))
        .group_by(IncomeCategory.name)
    )
    exp_by_cat = await db.execute(
        select(ExpenseCategory.name, func.sum(Expense.amount).label("total"))
        .join(Expense, Expense.category_id == ExpenseCategory.id)
        .where(year_filter(Expense, year))
        .group_by(ExpenseCategory.name)
    )
    inc_total = await db.execute(
        select(func.sum(Income.amount)).where(year_filter(Income, year))
    )
    exp_total = await db.execute(
        select(func.sum(Expense.amount)).where(year_filter(Expense, year))
    )

    dep_result = await db.execute(
        select(Depreciation).options(selectinload(Depreciation.expense))
    )
    depreciations = dep_result.scalars().all()
    dep_data = [calc_dep_overview(d) for d in depreciations]
    dep_this_year = sum(calc_depreciation_for_year(d, year) for d in depreciations)

    total_inc = inc_total.scalar() or 0
    total_exp = exp_total.scalar() or 0

    hours_by_cat = await db.execute(
        select(HourCategory.name, func.sum(TimeEntry.hours).label("total"))
        .join(TimeEntry, TimeEntry.category_id == HourCategory.id)
        .where(year_filter(TimeEntry, year))
        .group_by(HourCategory.name)
    )
    hours_total_result = await db.execute(
        select(func.sum(TimeEntry.hours)).where(year_filter(TimeEntry, year))
    )
    total_hours = hours_total_result.scalar() or 0.0

    return templates.TemplateResponse(request, "reports.html", {
        "year": year,
        "years": list(range(2022, datetime.now().year + 2)),
        "income_by_cat": [(r.name, float(r.total)) for r in inc_by_cat],
        "expense_by_cat": [(r.name, float(r.total)) for r in exp_by_cat],
        "total_income": total_inc,
        "total_expenses": total_exp,
        "profit": total_inc - total_exp,
        "depreciations": dep_data,
        "dep_this_year": dep_this_year,
        "hours_by_cat": [(r.name, float(r.total)) for r in hours_by_cat],
        "total_hours": total_hours,
        "urencriterium": URENCRITERIUM_UREN,
        "urencriterium_pct": min(100, round(total_hours / URENCRITERIUM_UREN * 100)) if URENCRITERIUM_UREN else 0,
    })


@router.get("/jaaroverzicht", response_class=HTMLResponse)
async def yearly_overview(request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user

    year = int(request.query_params.get("jaar", datetime.now().year))

    monthly_inc = await db.execute(
        select(
            func.strftime("%m", Income.date).label("month"),
            func.sum(Income.amount).label("total")
        ).where(year_filter(Income, year))
        .group_by(func.strftime("%m", Income.date))
    )
    monthly_exp = await db.execute(
        select(
            func.strftime("%m", Expense.date).label("month"),
            func.sum(Expense.amount).label("total")
        ).where(year_filter(Expense, year))
        .group_by(func.strftime("%m", Expense.date))
    )
    all_inc = await db.execute(
        select(Income).options(selectinload(Income.category))
        .where(year_filter(Income, year)).order_by(Income.date)
    )
    all_exp = await db.execute(
        select(Expense).options(selectinload(Expense.category))
        .where(year_filter(Expense, year)).order_by(Expense.date)
    )

    inc_by_month = {int(r.month): float(r.total) for r in monthly_inc}
    exp_by_month = {int(r.month): float(r.total) for r in monthly_exp}

    month_names = ["Januari", "Februari", "Maart", "April", "Mei", "Juni",
                   "Juli", "Augustus", "September", "Oktober", "November", "December"]

    rows = []
    cumulative_profit = 0
    for m in range(1, 13):
        inc = inc_by_month.get(m, 0)
        exp = exp_by_month.get(m, 0)
        profit = inc - exp
        cumulative_profit += profit
        rows.append({
            "month": m, "name": month_names[m - 1],
            "income": inc, "expense": exp,
            "profit": profit, "cumulative": cumulative_profit,
        })

    total_inc = sum(r["income"] for r in rows)
    total_exp = sum(r["expense"] for r in rows)

    inc_by_cat = await db.execute(
        select(IncomeCategory.name, func.sum(Income.amount).label("total"))
        .join(Income, Income.category_id == IncomeCategory.id)
        .where(year_filter(Income, year))
        .group_by(IncomeCategory.name)
    )
    exp_by_cat = await db.execute(
        select(ExpenseCategory.name, func.sum(Expense.amount).label("total"))
        .join(Expense, Expense.category_id == ExpenseCategory.id)
        .where(year_filter(Expense, year))
        .group_by(ExpenseCategory.name)
    )
    unpaid = await db.execute(
        select(Income).options(selectinload(Income.category))
        .where(Income.status == "niet_betaald", year_filter(Income, year))
        .order_by(Income.date)
    )

    open_items_result = await db.execute(select(ChecklistItem).where(ChecklistItem.status == "open"))
    open_checklist_items = open_items_result.scalars().all()

    closure_result = await db.execute(select(YearClosure).where(YearClosure.year == year))
    closure = closure_result.scalar_one_or_none()

    fy_result = await db.execute(select(FiscalYear).where(FiscalYear.year == year))
    fiscal_year = fy_result.scalar_one_or_none()

    return templates.TemplateResponse(request, "yearly.html", {
        "year": year,
        "years": list(range(2022, datetime.now().year + 2)),
        "rows": rows,
        "total_income": total_inc,
        "total_expenses": total_exp,
        "profit": total_inc - total_exp,
        "income_by_cat": [(r.name, float(r.total)) for r in inc_by_cat],
        "expense_by_cat": [(r.name, float(r.total)) for r in exp_by_cat],
        "unpaid_invoices": unpaid.scalars().all(),
        "all_incomes": all_inc.scalars().all(),
        "all_expenses": all_exp.scalars().all(),
        "chart_months": [r["name"][:3] for r in rows],
        "chart_income": [r["income"] for r in rows],
        "chart_expenses": [r["expense"] for r in rows],
        "chart_cumulative": [r["cumulative"] for r in rows],
        "open_checklist_items": open_checklist_items,
        "closure": closure,
        "fiscal_year": fiscal_year,
        "warn_open_items": request.query_params.get("warn") == "1",
    })


@router.post("/jaaroverzicht/afsluiten")
async def close_year(
    request: Request,
    jaar: int = Form(...),
    force: Optional[str] = Form(None),
    db: AsyncSession = Depends(get_db),
):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user

    open_items_result = await db.execute(select(ChecklistItem).where(ChecklistItem.status == "open"))
    open_items = open_items_result.scalars().all()

    if open_items and force != "on":
        return RedirectResponse(f"/jaaroverzicht?jaar={jaar}&warn=1", status_code=302)

    result = await db.execute(select(YearClosure).where(YearClosure.year == jaar))
    closure = result.scalar_one_or_none()
    if not closure:
        closure = YearClosure(year=jaar, open_items_at_closure=len(open_items))
        db.add(closure)
    else:
        closure.open_items_at_closure = len(open_items)
        closure.closed_at = datetime.now()
    await db.commit()
    return RedirectResponse(f"/jaaroverzicht?jaar={jaar}&closed=1", status_code=302)


@router.get("/export/inkomsten")
async def export_incomes(request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    result = await db.execute(
        select(Income).options(selectinload(Income.category)).order_by(Income.date.desc())
    )
    incomes = result.scalars().all()
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Factuurnummer", "Categorie", "Datum", "Bedrag", "Omschrijving", "Status"])
    for i in incomes:
        writer.writerow([i.invoice_number, i.category.name,
                         i.date.strftime("%d-%m-%Y"), f"{i.amount:.2f}",
                         i.description or "", i.status])
    output.seek(0)
    return StreamingResponse(io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=inkomsten.csv"})


@router.get("/export/uitgaven")
async def export_expenses(request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    result = await db.execute(
        select(Expense).options(selectinload(Expense.category)).order_by(Expense.date.desc())
    )
    expenses = result.scalars().all()
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Factuurnummer", "Categorie", "Datum", "Bedrag", "Omschrijving", "Afschrijving"])
    for e in expenses:
        writer.writerow([e.invoice_number, e.category.name,
                         e.date.strftime("%d-%m-%Y"), f"{e.amount:.2f}",
                         e.description or "", "Ja" if e.is_depreciable else "Nee"])
    output.seek(0)
    return StreamingResponse(io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=uitgaven.csv"})


@router.get("/export/jaaroverzicht")
async def export_yearly(request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    year = int(request.query_params.get("jaar", datetime.now().year))

    monthly_inc = await db.execute(
        select(func.strftime("%m", Income.date).label("month"), func.sum(Income.amount).label("total"))
        .where(year_filter(Income, year)).group_by(func.strftime("%m", Income.date))
    )
    monthly_exp = await db.execute(
        select(func.strftime("%m", Expense.date).label("month"), func.sum(Expense.amount).label("total"))
        .where(year_filter(Expense, year)).group_by(func.strftime("%m", Expense.date))
    )
    inc_by_month = {int(r.month): float(r.total) for r in monthly_inc}
    exp_by_month = {int(r.month): float(r.total) for r in monthly_exp}
    month_names = ["Januari","Februari","Maart","April","Mei","Juni",
                   "Juli","Augustus","September","Oktober","November","December"]

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Maand", "Inkomsten", "Uitgaven", "Winst/Verlies", "Cumulatief"])
    cumulative = 0
    for m in range(1, 13):
        inc = inc_by_month.get(m, 0)
        exp = exp_by_month.get(m, 0)
        profit = inc - exp
        cumulative += profit
        writer.writerow([month_names[m-1], f"{inc:.2f}", f"{exp:.2f}",
                         f"{profit:.2f}", f"{cumulative:.2f}"])
    output.seek(0)
    return StreamingResponse(io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=jaaroverzicht_{year}.csv"})


@router.get("/export/jaaroverzicht-xlsx")
async def export_yearly_xlsx(request: Request, db: AsyncSession = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    year = int(request.query_params.get("jaar", datetime.now().year))

    incomes_q = await db.execute(
        select(Income).options(selectinload(Income.category))
        .where(year_filter(Income, year)).order_by(Income.date)
    )
    incomes = incomes_q.scalars().all()

    expenses_q = await db.execute(
        select(Expense).options(selectinload(Expense.category))
        .where(year_filter(Expense, year)).order_by(Expense.date)
    )
    expenses = expenses_q.scalars().all()

    total_income = sum(i.amount for i in incomes)
    total_expenses = sum(e.amount for e in expenses)

    month_names = ["Januari","Februari","Maart","April","Mei","Juni",
                   "Juli","Augustus","September","Oktober","November","December"]

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="4A7C59")
    hdr_font = Font(color="FFFFFF", bold=True)

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # Sheet 1: Inkomsten
    ws1 = wb.active
    ws1.title = "Inkomsten"
    style_header(ws1, ["Factuurnummer", "Datum", "Categorie", "Bedrag (€)", "Status", "Omschrijving"])
    for i, inc in enumerate(incomes, 2):
        ws1.cell(i, 1, inc.invoice_number)
        ws1.cell(i, 2, inc.date.strftime("%d-%m-%Y"))
        ws1.cell(i, 3, inc.category.name if inc.category else "")
        ws1.cell(i, 4, inc.amount).number_format = '#,##0.00'
        ws1.cell(i, 5, inc.status)
        ws1.cell(i, 6, inc.description or "")
    ws1.cell(len(incomes) + 2, 3, "Totaal")
    ws1.cell(len(incomes) + 2, 3).font = Font(bold=True)
    ws1.cell(len(incomes) + 2, 4, total_income).number_format = '#,##0.00'
    ws1.cell(len(incomes) + 2, 4).font = Font(bold=True)
    auto_width(ws1)

    # Sheet 2: Uitgaven
    ws2 = wb.create_sheet("Uitgaven")
    style_header(ws2, ["Factuurnummer", "Datum", "Categorie", "Bedrag (€)", "Omschrijving"])
    for i, exp in enumerate(expenses, 2):
        ws2.cell(i, 1, exp.invoice_number)
        ws2.cell(i, 2, exp.date.strftime("%d-%m-%Y"))
        ws2.cell(i, 3, exp.category.name if exp.category else "")
        ws2.cell(i, 4, exp.amount).number_format = '#,##0.00'
        ws2.cell(i, 5, exp.description or "")
    ws2.cell(len(expenses) + 2, 3, "Totaal")
    ws2.cell(len(expenses) + 2, 3).font = Font(bold=True)
    ws2.cell(len(expenses) + 2, 4, total_expenses).number_format = '#,##0.00'
    ws2.cell(len(expenses) + 2, 4).font = Font(bold=True)
    auto_width(ws2)

    # Sheet 3: Resultaatoverzicht per maand
    ws3 = wb.create_sheet("Resultaatoverzicht")
    style_header(ws3, ["Maand", "Inkomsten (€)", "Uitgaven (€)", "Winst/Verlies (€)", "Cumulatief (€)"])
    inc_by_m = {}
    for i in incomes:
        inc_by_m[i.date.month] = inc_by_m.get(i.date.month, 0) + i.amount
    exp_by_m = {}
    for e in expenses:
        exp_by_m[e.date.month] = exp_by_m.get(e.date.month, 0) + e.amount
    cumulative = 0
    for m in range(1, 13):
        inc_m = inc_by_m.get(m, 0)
        exp_m = exp_by_m.get(m, 0)
        profit_m = inc_m - exp_m
        cumulative += profit_m
        row = m + 1
        ws3.cell(row, 1, month_names[m - 1])
        ws3.cell(row, 2, inc_m).number_format = '#,##0.00'
        ws3.cell(row, 3, exp_m).number_format = '#,##0.00'
        ws3.cell(row, 4, profit_m).number_format = '#,##0.00'
        ws3.cell(row, 5, cumulative).number_format = '#,##0.00'
    total_row = 14
    for col, val in [(1, "Totaal"), (2, total_income), (3, total_expenses),
                     (4, total_income - total_expenses)]:
        c = ws3.cell(total_row, col, val)
        c.font = Font(bold=True)
        if col > 1:
            c.number_format = '#,##0.00'
    auto_width(ws3)

    # Sheet 4: Per categorie
    ws4 = wb.create_sheet("Per categorie")
    style_header(ws4, ["Categorie", "Type", "Bedrag (€)", "Aantal"])
    inc_by_cat = {}
    for i in incomes:
        key = i.category.name if i.category else "Onbekend"
        inc_by_cat[key] = (inc_by_cat.get(key, (0, 0))[0] + i.amount,
                           inc_by_cat.get(key, (0, 0))[1] + 1)
    exp_by_cat = {}
    for e in expenses:
        key = e.category.name if e.category else "Onbekend"
        exp_by_cat[key] = (exp_by_cat.get(key, (0, 0))[0] + e.amount,
                           exp_by_cat.get(key, (0, 0))[1] + 1)
    row = 2
    for cat, (total, count) in sorted(inc_by_cat.items()):
        ws4.cell(row, 1, cat); ws4.cell(row, 2, "Inkomsten")
        ws4.cell(row, 3, total).number_format = '#,##0.00'
        ws4.cell(row, 4, count); row += 1
    for cat, (total, count) in sorted(exp_by_cat.items()):
        ws4.cell(row, 1, cat); ws4.cell(row, 2, "Uitgaven")
        ws4.cell(row, 3, total).number_format = '#,##0.00'
        ws4.cell(row, 4, count); row += 1
    auto_width(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=jaaroverzicht_{year}.xlsx"})


@router.get("/export/jaaroverzicht-pdf")
async def export_yearly_pdf(request: Request, db: AsyncSession = Depends(get_db)):
    user = await require_auth(request, db)
    if isinstance(user, RedirectResponse):
        return user
    year = int(request.query_params.get("jaar", datetime.now().year))

    monthly_inc = await db.execute(
        select(func.strftime("%m", Income.date).label("month"), func.sum(Income.amount).label("total"))
        .where(year_filter(Income, year)).group_by(func.strftime("%m", Income.date))
    )
    monthly_exp = await db.execute(
        select(func.strftime("%m", Expense.date).label("month"), func.sum(Expense.amount).label("total"))
        .where(year_filter(Expense, year)).group_by(func.strftime("%m", Expense.date))
    )
    inc_by_month = {int(r.month): float(r.total) for r in monthly_inc}
    exp_by_month = {int(r.month): float(r.total) for r in monthly_exp}

    fy_result = await db.execute(select(FiscalYear).where(FiscalYear.year == year))
    fiscal_year = fy_result.scalar_one_or_none()

    settings = request.state.settings
    company = settings.company_name if settings else ""
    buf = generate_yearly_pdf(year, inc_by_month, exp_by_month,
                              company_name=company, fiscal_year=fiscal_year)
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=jaaroverzicht_{year}.pdf"})
