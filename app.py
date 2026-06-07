import os
import re
import sqlite3
import uuid
import base64
import logging
import json
from datetime import datetime, date
from functools import wraps
from pathlib import Path

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, g, abort, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
import bleach

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

app = Flask(__name__)

secret_key = os.environ.get("SECRET_KEY", "")
if not secret_key or secret_key == "vervang_dit_met_een_lange_willekeurige_string":
    import secrets
    secret_key = secrets.token_hex(32)
    logger.warning("SECRET_KEY niet ingesteld — tijdelijke sleutel gegenereerd.")

app.secret_key = secret_key
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=False,
    PERMANENT_SESSION_LIFETIME=3600,
    MAX_CONTENT_LENGTH=10 * 1024 * 1024,
)

BASE_DIR    = Path(__file__).parent
DATA_DIR    = BASE_DIR / "data"
UPLOAD_BASE = BASE_DIR / "uploads"
DB_PATH     = DATA_DIR / "boekhouding.db"

DATA_DIR.mkdir(exist_ok=True)
UPLOAD_BASE.mkdir(exist_ok=True)

ALLOWED_EXTENSIONS = {"pdf", "jpg", "jpeg", "png"}
ALLOWED_CATEGORIES = [
    "Kantoor", "Marketing", "Transport", "Investering",
    "Vaste Lasten", "Gebruiksartikelen", "Overig"
]
for cat in ALLOWED_CATEGORIES:
    (UPLOAD_BASE / cat).mkdir(exist_ok=True)

# ── Database ──────────────────────────────────────────────────────────────────
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(str(DB_PATH), detect_types=sqlite3.PARSE_DECLTYPES)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(exc=None):
    db = g.pop("db", None)
    if db:
        db.close()

def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS gebruikers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gebruikersnaam TEXT UNIQUE NOT NULL,
            wachtwoord_hash TEXT NOT NULL,
            rol TEXT NOT NULL DEFAULT 'lezen',
            aangemaakt_op TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS transacties (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            datum DATE NOT NULL,
            factuurnummer TEXT,
            omschrijving TEXT NOT NULL,
            bedrag REAL NOT NULL CHECK(bedrag >= 0),
            type TEXT NOT NULL CHECK(type IN ('inkomst','uitgave')),
            categorie TEXT NOT NULL,
            bestand_pad TEXT,
            aangemaakt_op TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS activa (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            naam TEXT NOT NULL,
            aanschafdatum DATE NOT NULL,
            aanschafwaarde REAL NOT NULL CHECK(aanschafwaarde >= 0),
            aangemaakt_op TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS taken (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titel TEXT NOT NULL,
            omschrijving TEXT,
            verwacht_bedrag REAL DEFAULT 0,
            vervaldatum DATE,
            categorie TEXT,
            status TEXT NOT NULL DEFAULT 'gepland' CHECK(status IN ('gepland','in_uitvoering','voltooid')),
            prioriteit TEXT NOT NULL DEFAULT 'normaal' CHECK(prioriteit IN ('laag','normaal','hoog')),
            aangemaakt_op TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    # Rol-kolom toevoegen als migratie voor bestaande DB
    try:
        db.execute("ALTER TABLE gebruikers ADD COLUMN rol TEXT NOT NULL DEFAULT 'lezen'")
        db.commit()
    except Exception:
        pass

    existing = db.execute("SELECT id FROM gebruikers WHERE gebruikersnaam='admin'").fetchone()
    if not existing:
        db.execute(
            "INSERT INTO gebruikers (gebruikersnaam, wachtwoord_hash, rol) VALUES (?,?,?)",
            ("admin", generate_password_hash("admin123"), "admin")
        )
    else:
        db.execute("UPDATE gebruikers SET rol='admin' WHERE gebruikersnaam='admin'")
    db.commit()

# ── Hulpfuncties ──────────────────────────────────────────────────────────────
def login_vereist(f):
    @wraps(f)
    def dec(*a, **kw):
        if "gebruiker_id" not in session:
            return redirect(url_for("login"))
        return f(*a, **kw)
    return dec

def admin_vereist(f):
    @wraps(f)
    def dec(*a, **kw):
        if "gebruiker_id" not in session:
            return redirect(url_for("login"))
        if session.get("rol") != "admin":
            flash("Geen toegang — admin rechten vereist.", "error")
            return redirect(url_for("dashboard"))
        return f(*a, **kw)
    return dec

def schrijf_vereist(f):
    """Blokkeert lees-gebruikers voor schrijfacties."""
    @wraps(f)
    def dec(*a, **kw):
        if "gebruiker_id" not in session:
            return redirect(url_for("login"))
        if session.get("rol") == "lezen":
            flash("U heeft alleen leesrechten.", "error")
            return redirect(url_for("dashboard"))
        return f(*a, **kw)
    return dec

def sanitize(text, max_length=500):
    return bleach.clean(str(text), tags=[], strip=True)[:max_length]

def validate_bedrag(value):
    try:
        amount = float(str(value).replace(",", "."))
        if amount < 0 or amount > 1_000_000_000:
            raise ValueError
        return round(amount, 2)
    except (ValueError, TypeError):
        raise ValueError("Ongeldig bedrag")

def validate_datum(value):
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return value
    except ValueError:
        raise ValueError("Ongeldige datum")

def validate_categorie(cat):
    if cat not in ALLOWED_CATEGORIES:
        raise ValueError("Ongeldige categorie")
    return cat

def bon_bestandsnaam(factuurnummer: str, ext: str) -> str:
    """Gebruik factuurnummer als bestandsnaam (gesanitized), anders UUID."""
    if factuurnummer:
        veilig = re.sub(r"[^\w\-]", "_", factuurnummer.strip())[:60]
        if veilig:
            return f"{veilig}.{ext}"
    return f"{uuid.uuid4().hex}.{ext}"

def get_taken_open():
    """Aantal openstaande taken voor de sidebar-badge."""
    try:
        return get_db().execute(
            "SELECT COUNT(*) c FROM taken WHERE status != 'voltooid'"
        ).fetchone()["c"]
    except Exception:
        return 0

def bereken_activa(rijen):
    """20% lineaire afschrijving per jaar op aanschafwaarde, tot boekwaarde 0."""
    resultaat = []
    for a in rijen:
        aanschaf_datum = datetime.strptime(a["aanschafdatum"], "%Y-%m-%d").date()
        jaren_gebruikt = (date.today() - aanschaf_datum).days / 365.25
        jaarlijks = round(a["aanschafwaarde"] * 0.20, 2)
        max_afschrijving = a["aanschafwaarde"]
        afgeschreven = round(min(jaarlijks * jaren_gebruikt, max_afschrijving), 2)
        boekwaarde = round(max(a["aanschafwaarde"] - afgeschreven, 0), 2)
        jaren_resterend = round(max(0, 5 - jaren_gebruikt), 1)
        resultaat.append({
            "id": a["id"],
            "naam": a["naam"],
            "aanschafdatum": a["aanschafdatum"],
            "aanschafwaarde": a["aanschafwaarde"],
            "jaarlijkse_afschrijving": jaarlijks,
            "totaal_afgeschreven": afgeschreven,
            "boekwaarde": boekwaarde,
            "jaren_resterend": jaren_resterend,
        })
    return resultaat

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    if "gebruiker_id" in session:
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        gebruikersnaam = sanitize(request.form.get("gebruikersnaam",""), 100)
        wachtwoord = request.form.get("wachtwoord","")
        if not gebruikersnaam or not wachtwoord:
            flash("Vul alle velden in.", "error")
            return render_template("login.html")
        db = get_db()
        g_row = db.execute(
            "SELECT id, wachtwoord_hash, rol FROM gebruikers WHERE gebruikersnaam=?",
            (gebruikersnaam,)
        ).fetchone()
        if g_row and check_password_hash(g_row["wachtwoord_hash"], wachtwoord):
            session.clear()
            session["gebruiker_id"] = g_row["id"]
            session["gebruikersnaam"] = gebruikersnaam
            session["rol"] = g_row["rol"]
            session.permanent = True
            return redirect(url_for("dashboard"))
        flash("Ongeldige inloggegevens.", "error")
        logger.warning("Mislukte login: %s", gebruikersnaam)
    return render_template("login.html")

@app.route("/logout")
@login_vereist
def logout():
    session.clear()
    return redirect(url_for("login"))

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/")
@login_vereist
def dashboard():
    db = get_db()
    inkomsten = db.execute("SELECT COALESCE(SUM(bedrag),0) as t FROM transacties WHERE type='inkomst'").fetchone()["t"]
    uitgaven  = db.execute("SELECT COALESCE(SUM(bedrag),0) as t FROM transacties WHERE type='uitgave'").fetchone()["t"]

    activa_rijen = db.execute("SELECT * FROM activa").fetchall()
    activa_calc  = bereken_activa(activa_rijen)
    totaal_afschrijving = sum(a["totaal_afgeschreven"] for a in activa_calc)
    totaal_restwaarde   = sum(a["boekwaarde"] for a in activa_calc)

    recente = db.execute("SELECT * FROM transacties ORDER BY datum DESC, id DESC LIMIT 5").fetchall()

    huidig_jaar = str(date.today().year)
    maand_ink = {r["m"]: r["t"] for r in db.execute(
        "SELECT strftime('%m',datum) m, SUM(bedrag) t FROM transacties WHERE type='inkomst' AND strftime('%Y',datum)=? GROUP BY m", (huidig_jaar,)
    ).fetchall()}
    maand_uit = {r["m"]: r["t"] for r in db.execute(
        "SELECT strftime('%m',datum) m, SUM(bedrag) t FROM transacties WHERE type='uitgave' AND strftime('%Y',datum)=? GROUP BY m", (huidig_jaar,)
    ).fetchall()}

    ci = [round(maand_ink.get(f"{m:02d}", 0), 2) for m in range(1,13)]
    cu = [round(maand_uit.get(f"{m:02d}", 0), 2) for m in range(1,13)]
    cw = [round(ci[i]-cu[i], 2) for i in range(12)]

    cat_data = db.execute(
        "SELECT categorie, SUM(bedrag) t FROM transacties WHERE type='uitgave' AND strftime('%Y',datum)=? GROUP BY categorie ORDER BY t DESC",
        (huidig_jaar,)
    ).fetchall()

    taken_open = db.execute(
        "SELECT COUNT(*) c FROM taken WHERE status != 'voltooid'"
    ).fetchone()["c"]

    return render_template("index.html",
        pagina="dashboard",
        inkomsten=inkomsten, uitgaven=uitgaven, winst=inkomsten-uitgaven,
        totaal_afschrijving=round(totaal_afschrijving,2),
        totaal_restwaarde=round(totaal_restwaarde,2),
        recente_transacties=recente,
        huidig_jaar=huidig_jaar,
        chart_labels=json.dumps(["Jan","Feb","Mrt","Apr","Mei","Jun","Jul","Aug","Sep","Okt","Nov","Dec"]),
        chart_inkomsten=json.dumps(ci), chart_uitgaven=json.dumps(cu), chart_winst=json.dumps(cw),
        cat_labels=json.dumps([r["categorie"] for r in cat_data]),
        cat_totalen=json.dumps([round(r["t"],2) for r in cat_data]),
        taken_open=taken_open,
        gebruikersnaam=session.get("gebruikersnaam"), rol=session.get("rol"),
    )

# ── Transacties ───────────────────────────────────────────────────────────────
@app.route("/transacties")
@login_vereist
def transacties():
    db = get_db()
    rows = db.execute("SELECT * FROM transacties ORDER BY datum DESC, id DESC").fetchall()
    return render_template("index.html", pagina="transacties",
        transacties=rows, categorieen=ALLOWED_CATEGORIES,
        gebruikersnaam=session.get("gebruikersnaam"), rol=session.get("rol"),
        taken_open=get_taken_open())

@app.route("/transactie/toevoegen", methods=["POST"])
@schrijf_vereist
def transactie_toevoegen():
    try:
        datum          = validate_datum(request.form.get("datum",""))
        factuurnummer  = sanitize(request.form.get("factuurnummer",""), 50)
        omschrijving   = sanitize(request.form.get("omschrijving",""), 500)
        bedrag         = validate_bedrag(request.form.get("bedrag","0"))
        t_type         = request.form.get("type","")
        categorie      = validate_categorie(request.form.get("categorie",""))
        if t_type not in ("inkomst","uitgave"):
            flash("Ongeldig type.", "error"); return redirect(url_for("transacties"))
        if not omschrijving:
            flash("Omschrijving verplicht.", "error"); return redirect(url_for("transacties"))
        bestand_pad = None
        if "bon" in request.files:
            f = request.files["bon"]
            if f and f.filename and allowed_file(f.filename):
                ext = f.filename.rsplit(".",1)[1].lower()
                naam = bon_bestandsnaam(factuurnummer, ext)
                f.save(str(UPLOAD_BASE / categorie / naam))
                bestand_pad = f"{categorie}/{naam}"
        get_db().execute(
            "INSERT INTO transacties (datum,factuurnummer,omschrijving,bedrag,type,categorie,bestand_pad) VALUES (?,?,?,?,?,?,?)",
            (datum, factuurnummer, omschrijving, bedrag, t_type, categorie, bestand_pad)
        )
        get_db().commit()
        flash("Transactie toegevoegd.", "success")
    except ValueError as e:
        flash(f"Fout: {e}", "error")
    return redirect(url_for("transacties"))

@app.route("/transactie/bewerken/<int:tid>", methods=["GET","POST"])
@schrijf_vereist
def transactie_bewerken(tid):
    db = get_db()
    t = db.execute("SELECT * FROM transacties WHERE id=?", (tid,)).fetchone()
    if not t: abort(404)
    if request.method == "POST":
        try:
            datum         = validate_datum(request.form.get("datum",""))
            factuurnummer = sanitize(request.form.get("factuurnummer",""), 50)
            omschrijving  = sanitize(request.form.get("omschrijving",""), 500)
            bedrag        = validate_bedrag(request.form.get("bedrag","0"))
            t_type        = request.form.get("type","")
            categorie     = validate_categorie(request.form.get("categorie",""))
            if t_type not in ("inkomst","uitgave"):
                flash("Ongeldig type.","error"); return redirect(url_for("transactie_bewerken",tid=tid))
            bestand_pad = t["bestand_pad"]
            if "bon" in request.files:
                f = request.files["bon"]
                if f and f.filename and allowed_file(f.filename):
                    if bestand_pad:
                        p = UPLOAD_BASE / bestand_pad
                        if p.exists(): p.unlink()
                    ext  = f.filename.rsplit(".",1)[1].lower()
                    naam = bon_bestandsnaam(factuurnummer, ext)
                    f.save(str(UPLOAD_BASE / categorie / naam))
                    bestand_pad = f"{categorie}/{naam}"
            db.execute(
                "UPDATE transacties SET datum=?,factuurnummer=?,omschrijving=?,bedrag=?,type=?,categorie=?,bestand_pad=? WHERE id=?",
                (datum,factuurnummer,omschrijving,bedrag,t_type,categorie,bestand_pad,tid)
            )
            db.commit()
            flash("Transactie bijgewerkt.","success")
            return redirect(url_for("transacties"))
        except ValueError as e:
            flash(f"Fout: {e}","error")
    return render_template("index.html", pagina="transactie_bewerken",
        transactie=t, categorieen=ALLOWED_CATEGORIES,
        gebruikersnaam=session.get("gebruikersnaam"), rol=session.get("rol"),
        taken_open=get_taken_open())

@app.route("/transactie/verwijderen/<int:tid>", methods=["POST"])
@schrijf_vereist
def transactie_verwijderen(tid):
    db = get_db()
    rij = db.execute("SELECT bestand_pad FROM transacties WHERE id=?", (tid,)).fetchone()
    if rij and rij["bestand_pad"]:
        p = UPLOAD_BASE / rij["bestand_pad"]
        if p.exists(): p.unlink()
    db.execute("DELETE FROM transacties WHERE id=?", (tid,))
    db.commit()
    flash("Verwijderd.","success")
    return redirect(url_for("transacties"))

# ── Activa (vereenvoudigd: 20% per jaar, 5 jaar) ─────────────────────────────
@app.route("/activa")
@login_vereist
def activa():
    db = get_db()
    rijen = db.execute("SELECT * FROM activa ORDER BY aanschafdatum DESC").fetchall()
    return render_template("index.html", pagina="activa",
        activa=bereken_activa(rijen),
        gebruikersnaam=session.get("gebruikersnaam"), rol=session.get("rol"),
        taken_open=get_taken_open())

@app.route("/activum/toevoegen", methods=["POST"])
@schrijf_vereist
def activum_toevoegen():
    try:
        naam          = sanitize(request.form.get("naam",""), 200)
        aanschafdatum = validate_datum(request.form.get("aanschafdatum",""))
        aanschafwaarde= validate_bedrag(request.form.get("aanschafwaarde","0"))
        if not naam:
            flash("Naam verplicht.","error"); return redirect(url_for("activa"))
        get_db().execute(
            "INSERT INTO activa (naam,aanschafdatum,aanschafwaarde) VALUES (?,?,?)",
            (naam, aanschafdatum, aanschafwaarde)
        )
        get_db().commit()
        flash("Activum toegevoegd.","success")
    except ValueError as e:
        flash(f"Fout: {e}","error")
    return redirect(url_for("activa"))

@app.route("/activum/verwijderen/<int:aid>", methods=["POST"])
@schrijf_vereist
def activum_verwijderen(aid):
    get_db().execute("DELETE FROM activa WHERE id=?", (aid,))
    get_db().commit()
    flash("Activum verwijderd.","success")
    return redirect(url_for("activa"))

# ── Winst & Verlies ───────────────────────────────────────────────────────────
@app.route("/winst-verlies")
@login_vereist
def winst_verlies():
    db = get_db()
    jaar = request.args.get("jaar", str(date.today().year))
    try:
        jaar_int = int(jaar)
        if not (2000 <= jaar_int <= 2100): raise ValueError
    except ValueError:
        jaar_int = date.today().year; jaar = str(jaar_int)

    ink_rows = db.execute(
        "SELECT categorie, SUM(bedrag) totaal FROM transacties WHERE type='inkomst' AND strftime('%Y',datum)=? GROUP BY categorie", (jaar,)
    ).fetchall()
    uit_rows = db.execute(
        "SELECT categorie, SUM(bedrag) totaal FROM transacties WHERE type='uitgave' AND strftime('%Y',datum)=? GROUP BY categorie", (jaar,)
    ).fetchall()

    totaal_ink = sum(r["totaal"] for r in ink_rows)
    totaal_uit = sum(r["totaal"] for r in uit_rows)

    activa_rijen = db.execute("SELECT * FROM activa").fetchall()
    afschr_jaar = 0.0
    for a in activa_rijen:
        aanschaf = datetime.strptime(a["aanschafdatum"], "%Y-%m-%d").date()
        if aanschaf.year <= jaar_int <= aanschaf.year + 4:
            afschr_jaar += a["aanschafwaarde"] * 0.20

    winst = totaal_ink - totaal_uit - afschr_jaar
    jaren = list(range(date.today().year - 5, date.today().year + 1))

    return render_template("index.html", pagina="winst_verlies",
        inkomsten_rows=ink_rows, uitgaven_rows=uit_rows,
        totaal_inkomsten=totaal_ink, totaal_uitgaven=totaal_uit,
        afschrijvingen_jaar=round(afschr_jaar,2), winst=round(winst,2),
        geselecteerd_jaar=jaar_int, jaren=jaren,
        gebruikersnaam=session.get("gebruikersnaam"), rol=session.get("rol"),
        taken_open=get_taken_open())

# ── OCR / Bon scannen ─────────────────────────────────────────────────────────
@app.route("/ocr", methods=["POST"])
@login_vereist
def ocr():
    api_key = os.environ.get("OPENAI_API_KEY","")
    if not api_key or api_key.startswith("your_"):
        return jsonify({"error":"OpenAI API key niet ingesteld."}), 503
    if "bestand" not in request.files:
        return jsonify({"error":"Geen bestand."}), 400
    bestand = request.files["bestand"]
    if not bestand or not allowed_file(bestand.filename):
        return jsonify({"error":"Ongeldig bestandstype."}), 400
    try:
        data = bestand.read(5*1024*1024)
        b64  = base64.b64encode(data).decode()
        ext  = bestand.filename.rsplit(".",1)[1].lower()
        media_type = "application/pdf" if ext=="pdf" else f"image/{ext}"

        from openai import OpenAI
        client = OpenAI(api_key=api_key)
        prompt = (
            "Analyseer dit bon/factuur en retourneer UITSLUITEND geldig JSON (geen markdown, geen uitleg):\n"
            '{"leverancier":"...","factuurnummer":"...","datum":"YYYY-MM-DD of leeg",'
            '"btw":0,"totaalbedrag":0,"categorie":"Kantoor|Marketing|Transport|Investering|Vaste Lasten|Gebruiksartikelen|Overig",'
            '"omschrijving":"korte omschrijving","type":"inkomst of uitgave"}'
        )
        msgs_content = [{"type":"text","text":prompt}]
        if ext != "pdf":
            msgs_content.append({"type":"image_url","image_url":{"url":f"data:{media_type};base64,{b64}"}})
        else:
            msgs_content.append({"type":"text","text":f"[PDF base64 inhoud beschikbaar, {len(b64)} tekens]"})

        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"user","content":msgs_content}],
            max_tokens=500,
        )
        tekst = resp.choices[0].message.content.strip()
        tekst = re.sub(r"```(?:json)?|```","",tekst).strip()
        r = json.loads(tekst)
        veilig = {
            "leverancier":  sanitize(str(r.get("leverancier","")), 200),
            "factuurnummer":sanitize(str(r.get("factuurnummer","")), 50),
            "datum":        r.get("datum",""),
            "btw":          float(r.get("btw",0) or 0),
            "totaalbedrag": float(r.get("totaalbedrag",0) or 0),
            "categorie":    r.get("categorie","Overig") if r.get("categorie") in ALLOWED_CATEGORIES else "Overig",
            "omschrijving": sanitize(str(r.get("omschrijving","")), 500),
            "type":         r.get("type","uitgave") if r.get("type") in ("inkomst","uitgave") else "uitgave",
        }
        return jsonify(veilig)
    except Exception as e:
        logger.error("OCR fout: %s", e)
        return jsonify({"error":"OCR mislukt."}), 500

# ── Taken & Geplande Kosten ───────────────────────────────────────────────────
@app.route("/taken")
@login_vereist
def taken():
    db = get_db()
    rows = db.execute("SELECT * FROM taken ORDER BY CASE prioriteit WHEN 'hoog' THEN 1 WHEN 'normaal' THEN 2 ELSE 3 END, vervaldatum ASC NULLS LAST").fetchall()
    return render_template("index.html", pagina="taken",
        taken=rows, categorieen=ALLOWED_CATEGORIES,
        gebruikersnaam=session.get("gebruikersnaam"), rol=session.get("rol"),
        taken_open=get_taken_open())

@app.route("/taak/toevoegen", methods=["POST"])
@schrijf_vereist
def taak_toevoegen():
    try:
        titel           = sanitize(request.form.get("titel",""), 200)
        omschrijving    = sanitize(request.form.get("omschrijving",""), 1000)
        verwacht_bedrag = validate_bedrag(request.form.get("verwacht_bedrag","0") or "0")
        vervaldatum     = request.form.get("vervaldatum","") or None
        if vervaldatum: vervaldatum = validate_datum(vervaldatum)
        categorie       = request.form.get("categorie","Overig")
        if categorie not in ALLOWED_CATEGORIES: categorie = "Overig"
        prioriteit      = request.form.get("prioriteit","normaal")
        if prioriteit not in ("laag","normaal","hoog"): prioriteit = "normaal"
        if not titel:
            flash("Titel verplicht.","error"); return redirect(url_for("taken"))
        get_db().execute(
            "INSERT INTO taken (titel,omschrijving,verwacht_bedrag,vervaldatum,categorie,prioriteit) VALUES (?,?,?,?,?,?)",
            (titel, omschrijving, verwacht_bedrag, vervaldatum, categorie, prioriteit)
        )
        get_db().commit()
        flash("Taak toegevoegd.","success")
    except ValueError as e:
        flash(f"Fout: {e}","error")
    return redirect(url_for("taken"))

@app.route("/taak/status/<int:tid>", methods=["POST"])
@schrijf_vereist
def taak_status(tid):
    status = request.form.get("status","gepland")
    if status not in ("gepland","in_uitvoering","voltooid"):
        abort(400)
    get_db().execute("UPDATE taken SET status=? WHERE id=?", (status, tid))
    get_db().commit()
    return redirect(url_for("taken"))

@app.route("/taak/verwijderen/<int:tid>", methods=["POST"])
@schrijf_vereist
def taak_verwijderen(tid):
    get_db().execute("DELETE FROM taken WHERE id=?", (tid,))
    get_db().commit()
    flash("Taak verwijderd.","success")
    return redirect(url_for("taken"))

# ── Excel Export ──────────────────────────────────────────────────────────────
@app.route("/export/excel")
@login_vereist
def export_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    wb = Workbook()
    H_FILL = PatternFill("solid", start_color="1A3A6B")
    H_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    N_FONT = Font(name="Calibri", size=10)
    THIN   = Side(style="thin", color="E5E7EB")
    BRD    = Border(bottom=THIN)

    def header(ws, hdrs, widths):
        ws.append(hdrs)
        for i,c in enumerate(ws[1],1):
            c.font=H_FONT; c.fill=H_FILL
            c.alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height=22
        for i,w in enumerate(widths,1):
            ws.column_dimensions[get_column_letter(i)].width=w

    def norm_rows(ws):
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font=N_FONT; c.border=BRD
                c.alignment=Alignment(vertical="center")
            ws.row_dimensions[c.row].height=17

    # Transacties
    ws1 = wb.active; ws1.title="Transacties"
    header(ws1,["ID","Datum","Factuurnr","Omschrijving","Type","Categorie","Bedrag €","Bijlage"],
               [5,12,16,40,12,16,12,8])
    INK=PatternFill("solid",start_color="F0FDF4"); UIT=PatternFill("solid",start_color="FEF2F2")
    for t in db.execute("SELECT * FROM transacties ORDER BY datum DESC").fetchall():
        r=ws1.max_row+1
        ws1.append([t["id"],t["datum"],t["factuurnummer"] or "",t["omschrijving"],
                    t["type"].capitalize(),t["categorie"],t["bedrag"],"Ja" if t["bestand_pad"] else "Nee"])
        fill=INK if t["type"]=="inkomst" else UIT
        for c in ws1[r]: c.fill=fill
    ws1.freeze_panes="A2"; norm_rows(ws1)

    # Activa
    ws2 = wb.create_sheet("Activa"); 
    header(ws2,["ID","Naam","Aanschafdatum","Aanschafwaarde €","Jaarl. 20% €","Afgeschreven €","Boekwaarde €","Resterend jr"],
               [5,24,14,18,14,16,14,12])
    for a in bereken_activa(db.execute("SELECT * FROM activa ORDER BY aanschafdatum DESC").fetchall()):
        ws2.append([a["id"],a["naam"],a["aanschafdatum"],a["aanschafwaarde"],
                    a["jaarlijkse_afschrijving"],a["totaal_afgeschreven"],a["boekwaarde"],a["jaren_resterend"]])
    ws2.freeze_panes="A2"; norm_rows(ws2)

    # Taken
    ws3 = wb.create_sheet("Taken")
    header(ws3,["ID","Titel","Categorie","Verwacht €","Vervaldatum","Prioriteit","Status"],
               [5,30,16,12,14,12,14])
    for t in db.execute("SELECT * FROM taken ORDER BY vervaldatum").fetchall():
        ws3.append([t["id"],t["titel"],t["categorie"] or "",t["verwacht_bedrag"] or 0,
                    t["vervaldatum"] or "",t["prioriteit"],t["status"]])
    ws3.freeze_panes="A2"; norm_rows(ws3)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"boekhouding_{date.today().isoformat()}.xlsx")

# ── Backup ────────────────────────────────────────────────────────────────────
@app.route("/backup/maken", methods=["POST"])
@admin_vereist
def backup_maken():
    import zipfile
    backup_dir = BASE_DIR / "backups"
    backup_dir.mkdir(exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    pad = backup_dir / f"backup_{ts}.zip"
    with zipfile.ZipFile(str(pad),"w",zipfile.ZIP_DEFLATED) as zf:
        if DB_PATH.exists(): zf.write(str(DB_PATH),"boekhouding.db")
        for f in UPLOAD_BASE.rglob("*"):
            if f.is_file(): zf.write(str(f), str(f.relative_to(BASE_DIR)))
    flash(f"Backup aangemaakt: {pad.name}","success")
    logger.info("Backup: %s door %s", pad.name, session.get("gebruikersnaam"))
    return redirect(url_for("instellingen"))

@app.route("/backup/downloaden/<naam>")
@admin_vereist
def backup_downloaden(naam):
    if not re.fullmatch(r"backup_\d{8}_\d{6}\.zip", naam): abort(400)
    pad = BASE_DIR / "backups" / naam
    if not pad.exists(): abort(404)
    return send_file(str(pad), as_attachment=True, download_name=naam)

@app.route("/backup/verwijderen/<naam>", methods=["POST"])
@admin_vereist
def backup_verwijderen(naam):
    if not re.fullmatch(r"backup_\d{8}_\d{6}\.zip", naam): abort(400)
    pad = BASE_DIR / "backups" / naam
    if pad.exists(): pad.unlink()
    flash(f"Backup verwijderd: {naam}","success")
    return redirect(url_for("instellingen"))

# ── Instellingen ──────────────────────────────────────────────────────────────
@app.route("/instellingen")
@login_vereist
def instellingen():
    db = get_db()
    backups = sorted((BASE_DIR/"backups").glob("backup_*.zip"),
                     key=lambda p: p.stat().st_mtime, reverse=True) \
              if (BASE_DIR/"backups").exists() else []
    gebruikers = db.execute("SELECT id, gebruikersnaam, rol, aangemaakt_op FROM gebruikers ORDER BY id").fetchall() \
                 if session.get("rol")=="admin" else []
    return render_template("index.html", pagina="instellingen",
        backups=[p.name for p in backups[:20]],
        gebruikers=gebruikers,
        gebruikersnaam=session.get("gebruikersnaam"), rol=session.get("rol"),
        taken_open=get_taken_open())

@app.route("/instellingen/wachtwoord", methods=["POST"])
@login_vereist
def wachtwoord_wijzigen():
    huidig  = request.form.get("huidig_wachtwoord","")
    nieuw   = request.form.get("nieuw_wachtwoord","")
    bevestig= request.form.get("bevestig_wachtwoord","")
    if not all([huidig,nieuw,bevestig]):
        flash("Vul alle velden in.","error"); return redirect(url_for("instellingen"))
    if len(nieuw) < 8:
        flash("Minimaal 8 tekens vereist.","error"); return redirect(url_for("instellingen"))
    if nieuw != bevestig:
        flash("Wachtwoorden komen niet overeen.","error"); return redirect(url_for("instellingen"))
    db = get_db()
    g_row = db.execute("SELECT wachtwoord_hash FROM gebruikers WHERE id=?", (session["gebruiker_id"],)).fetchone()
    if not g_row or not check_password_hash(g_row["wachtwoord_hash"], huidig):
        flash("Huidig wachtwoord onjuist.","error"); return redirect(url_for("instellingen"))
    db.execute("UPDATE gebruikers SET wachtwoord_hash=? WHERE id=?",
               (generate_password_hash(nieuw), session["gebruiker_id"]))
    db.commit()
    flash("Wachtwoord gewijzigd.","success")
    return redirect(url_for("instellingen"))

@app.route("/instellingen/gebruiker/toevoegen", methods=["POST"])
@admin_vereist
def gebruiker_toevoegen():
    naam  = sanitize(request.form.get("gebruikersnaam",""), 100)
    pw    = request.form.get("wachtwoord","")
    rol   = request.form.get("rol","lezen")
    if rol not in ("admin","lezen"): rol="lezen"
    if not naam or not pw:
        flash("Naam en wachtwoord verplicht.","error"); return redirect(url_for("instellingen"))
    if len(pw) < 8:
        flash("Minimaal 8 tekens vereist.","error"); return redirect(url_for("instellingen"))
    try:
        get_db().execute("INSERT INTO gebruikers (gebruikersnaam,wachtwoord_hash,rol) VALUES (?,?,?)",
                        (naam, generate_password_hash(pw), rol))
        get_db().commit()
        flash(f"Gebruiker '{naam}' aangemaakt.","success")
    except Exception:
        flash("Gebruikersnaam al in gebruik.","error")
    return redirect(url_for("instellingen"))

@app.route("/instellingen/gebruiker/verwijderen/<int:uid>", methods=["POST"])
@admin_vereist
def gebruiker_verwijderen(uid):
    if uid == session["gebruiker_id"]:
        flash("U kunt uw eigen account niet verwijderen.","error")
        return redirect(url_for("instellingen"))
    get_db().execute("DELETE FROM gebruikers WHERE id=?", (uid,))
    get_db().commit()
    flash("Gebruiker verwijderd.","success")
    return redirect(url_for("instellingen"))

# ── Start ─────────────────────────────────────────────────────────────────────
with app.app_context():
    init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
